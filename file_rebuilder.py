import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from docx import Document
import PyPDF2

# Define a function to process the Excel files
def process_excel(file_path, worksheet):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    max_row = sheet.max_row
    max_column = sheet.max_column
    for row in range(1, max_row+1):
        for col in range(1, max_column+1):
            cell_value = sheet.cell(row=row, column=col).value
            worksheet.cell(row=row, column=1, value=file_path)
            worksheet.cell(row=row, column=2, value=cell_value)

# Define a function to process the Word files
def process_word(file_path, worksheet):
    document = Document(file_path)
    for paragraph in document.paragraphs:
        text = paragraph.text
        sentences = text.split('.')
        for sentence in sentences:
            worksheet.cell(row=worksheet.max_row+1, column=1, value=file_path)
            worksheet.cell(row=worksheet.max_row, column=2, value=sentence)

# Define a function to process the PDF files
def process_pdf(file_path, worksheet):
    pdf_file = open(file_path, 'rb')
    pdf_reader = PyPDF2.PdfFileReader(pdf_file)
    num_pages = pdf_reader.numPages
    for page_num in range(0, num_pages):
        page = pdf_reader.getPage(page_num)
        text = page.extractText()
        sentences = text.split('.')
        for sentence in sentences:
            worksheet.cell(row=worksheet.max_row+1, column=1, value=file_path)
            worksheet.cell(row=worksheet.max_row, column=2, value=sentence)

# Create a new Excel workbook and add a worksheet
wb = openpyxl.Workbook()
worksheet = wb.active
worksheet.cell(row=1, column=1, value="Original File")
worksheet.cell(row=1, column=2, value="Text")
worksheet.cell(row=1, column=1).font = Font(bold=True)
worksheet.cell(row=1, column=2).font = Font(bold=True)

# Set the column widths
worksheet.column_dimensions[get_column_letter(1)].width = 40
worksheet.column_dimensions[get_column_letter(2)].width = 80

# Get the current working directory
cwd = os.getcwd()

# Process all the Excel, Word, and PDF
